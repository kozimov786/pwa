import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .i18n import t

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
ALT_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _build_excel(meta: list[tuple[str, object]], headers: list[str], rows: list[list], lang: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    ws.merge_cells("A1:D1")
    ws["A1"] = "GOKLE"
    ws["A1"].font = Font(bold=True, size=16)

    row_idx = 2
    for label, value in meta:
        ws[f"A{row_idx}"] = label
        ws[f"A{row_idx}"].font = Font(bold=True)
        ws[f"B{row_idx}"] = value
        row_idx += 1

    header_row = row_idx + 1
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for i, values in enumerate(rows):
        row = header_row + 1 + i
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            if col == 3:
                cell.number_format = '0.00"$"'
            if col == 4:
                cell.number_format = '#,##0.00"$"'
            if i % 2 == 1:
                cell.fill = ALT_FILL

    widths = [30, 16, 18, 20]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_quotation_excel(product_name: str, weight_kg: float, legs: list[dict], lang: str = "en") -> bytes:
    meta = [
        (t(lang, "product"), product_name),
        (t(lang, "date"), date.today().isoformat()),
        (t(lang, "weight"), weight_kg),
    ]
    headers = [t(lang, "col_destination"), t(lang, "col_total_kg"), t(lang, "col_price_per_kg"), t(lang, "col_total_cost")]
    rows = [[leg["destination"], weight_kg, leg["price_per_kg_usd"], leg["total_usd"]] for leg in legs]
    return _build_excel(meta, headers, rows, lang)


def generate_comparison_excel(destination: str, rows: list[dict], lang: str = "en") -> bytes:
    meta = [
        (t(lang, "destination"), destination),
        (t(lang, "date"), date.today().isoformat()),
    ]
    headers = [t(lang, "col_product"), t(lang, "col_total_kg"), t(lang, "col_price_per_kg"), t(lang, "col_total_cost")]
    table_rows = [[row["product_name"], row["weight_kg"], row["price_per_kg_usd"], row["total_usd"]] for row in rows]
    return _build_excel(meta, headers, table_rows, lang)
